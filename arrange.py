import openpyxl
import sys
import json
import re

# ACCEPT JSON DATA FROM COMMAND-LINE ARGUMENTS
data_json = sys.argv[1]
data = json.loads(data_json)

# TAKING DATA FROM DATA DICTIONARY
rooms = list()
details = list()
date = data['date']
time = data['time']
rooms = data['rooms']
details = data['details']
sem = details[0].get("sem")

print(rooms)

# Extract room numbers and capacities
room_numbers = [room['room_no'] for room in rooms]
capacity = [room['capacity'] for room in rooms]

# Load the existing workbook and get the sheet containing the data
input_file = '.\\uploadedExcels\\S'+str(sem)+'.xlsx'  # Replace with your actual file path
wb = openpyxl.load_workbook(input_file)
ws = wb.active  # Assuming the data is in the active sheet

# Read the column with the names and alphanumeric codes (assuming it's in the first column)
data = []
for row in ws.iter_rows(min_col=1, max_col=1, min_row=1):
    for cell in row:
        data.append(cell.value)

# Function to separate name and alphanumeric code
def separate_name_and_code(value):
    match = re.search(r"(.*)\s(1RG\S+)", value)
    if match:
        name = match.group(1).strip()  # Everything before "1RG"
        code = match.group(2).strip()  # "1RG" and what follows
    else:
        name = value.strip()  # Default to full value if no match
        code = ""  # No code found
    return name, code

# Create a new workbook for the seating arrangement
new_wb = openpyxl.Workbook()

# Iterate over the rooms and assign data based on their capacities
data_index = 0
for i, room in enumerate(room_numbers):
    if data_index >= len(data):
        break  # Stop if no more data to allocate
    
    # Get the capacity for the current room
    room_capacity = capacity[i]
    
    # Create or select the sheet for the room
    if i == 0:
        sheet = new_wb.active
        sheet.title = f"Room {room}"
    else:
        sheet = new_wb.create_sheet(f"Room {room}")
    
    # Write the headers
    sheet['A1'] = "Name"
    sheet['B1'] = "USN"
    
    # Get the data chunk for the room
    chunk = data[data_index:data_index + room_capacity]
    data_index += room_capacity  # Update the index
    
    # Write the names and alphanumeric codes into the sheet
    for row_num, item in enumerate(chunk, start=2):
        name, code = separate_name_and_code(item)
        sheet[f'A{row_num}'] = name
        sheet[f'B{row_num}'] = code
    
    # Adjust column widths
    sheet.column_dimensions['A'].width = 35  # Adjust width for Name column
    sheet.column_dimensions['B'].width = 20  # Adjust width for Register Number column
    
    # Optionally, auto-adjust row heights for better readability
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="center")

# Save the new workbook
new_file = '.\\arrangementExcels\\'+'S'+str(sem)+'_'+date+'_'+time+'.xlsx'  # Name of the new file
new_wb.save(new_file)

print(f"Seating arrangement saved in {new_file}")
